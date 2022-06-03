import openpyxl

wb = openpyxl.load_workbook("new.xlsx")

ws1 = wb["Expenses"]


def create_sheets():
    ws1 = wb.create_sheet(title='Expenses')

    ws2 = wb.create_sheet(index=0, title='Purchases')


print(wb.sheetnames)


def delete_sheet(name):
    del wb[name]


# delete_sheet('Expenses1')

def change_data_in_cell():
    print(ws1["A1"])
    print(ws1["A1"].value)
    ws1['A1'] = 'Items'


change_data_in_cell()


wb.save("new.xlsx")
