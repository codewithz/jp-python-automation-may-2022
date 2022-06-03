from codecs import backslashreplace_errors
import openpyxl

from openpyxl.styles import Font

wb = openpyxl.load_workbook("balances.xlsx")

ws = wb["Sheet1"]

ws["B12"] = "=SUM(B2:B9)"

ws["B13"] = "=AVERAGE(B2:B9)"

wb.save("balances.xlsx")

# Balance after one year

ws["D1"] = "Balance after a year"
ws['D1'].font = Font(bold=True, name='Arial', size=12)

# Balance after a year = (Balance* Interest Rate)+Balance

for i in range(2, 10):
    balance = ws.cell(row=i, column=2).value
    print("Balance:", balance)
    interest = ws.cell(row=i, column=3).value
    print("Interest", interest)
    final_balance = (balance*interest)+balance
    ws.cell(row=i, column=4).value = final_balance


wb.save("balances.xlsx")
