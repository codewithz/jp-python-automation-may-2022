import openpyxl

wb = openpyxl.load_workbook("myfile.xlsx")
ws = wb['Sheet1']


def merge_cells():
    ws.merge_cells('B2:C9')
    ws['B2'] = 'Merged Data'
    wb.save('myfile.xlsx')


def unmerge_cells():
    ws.unmerge_cells('B2:C9')
    wb.save('myfile.xlsx')


# merge_cells()
unmerge_cells()
