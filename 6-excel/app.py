import openpyxl

# wb=openpyxl.Workbook()

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb)
print(wb.sheetnames)

sheet = wb["Sheet1"]
wb.create_sheet("Sheet2", 1)

print("Sheet Created")

cell = sheet["a1"]
print(cell.value)

print("Row:", cell.row)
print("Col:", cell.column)
print("Coordinate:", cell.coordinate)

print("Max Row:", sheet.max_row)
print("Max Column:", sheet.max_column)
print(40*'-')
for row in range(1, sheet.max_row+1):
    for column in range(1, sheet.max_column+1):
        cell = sheet.cell(row, column)
        print("Row:", row, " | Column:", column)
        print(cell.value)

print(40*'-')

column = sheet["a"]
print("Column A", column)
# print(dir(column))
print(40*'-')
cells = sheet["a:c"]
print("A:C", cells)
print(40*'-')
print(sheet[1:3])

sheet.append([1010, 34, 32.34])

wb.save("transactions.xlsx")
